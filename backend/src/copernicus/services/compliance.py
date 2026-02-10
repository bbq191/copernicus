"""合规审核服务

基于 CSV/XLSX 规则 + ASR 转写文本，使用 Map-Reduce 策略进行 LLM 逐段合规判定。
参照 EvaluatorService 的架构，确保每次 LLM 调用 num_ctx 控制在 8192，显存稳定。

流程：
  parse_rules(file_bytes) -> rules[]
  audit(rules, transcript_entries) ->
    按 entry 分组 -> Map 并发审核各 chunk -> Reduce 生成摘要 -> ComplianceReport
"""

import asyncio
import csv
import io
import json
import logging
import re
from collections.abc import Iterable

from copernicus.config import Settings
from copernicus.exceptions import ComplianceError
from copernicus.schemas.compliance import ComplianceReport, ComplianceRule, Violation
from copernicus.services.llm import OllamaClient
from copernicus.utils.llm_parse import extract_json_array, strip_think_tags
from copernicus.utils.types import ProgressCallback

logger = logging.getLogger(__name__)

_AUDIT_SYSTEM_PROMPT = """\
你是一个保险行业合规审核专家，执行严格的合规质检任务。

### 核心工作方法
1. 你必须逐条对照【审核标准】中的每一条规则，检查【语音转录文本】中是否存在违反。
2. 宁严勿松：有疑似违规的内容也必须报告（标记为 medium），绝不放过。
3. 对于包含"不允许出现"、"不得"、"禁止"等关键词的规则，执行精确匹配——只要转录文本中出现了规则禁止的字样或语义相近的表述，即判定为违规。
4. ASR 转写存在同音字误差（如"保种"可能是"保证"），你必须结合上下文语义判断，不要因为同音字差异而漏判。

### 绝对格式约束
1. 你必须且只能输出一段合法的 JSON 数组。
2. 严禁输出任何 Markdown 标记、开场白、结束语或解释文字。
3. 如果没有发现违规，输出空数组 []。

### JSON 输出结构（数组中的每个元素）
{
    "rule_id": 对应审核标准的编号(整数),
    "timestamp": "违规发生的时间(来自转录文本中的时间标记，如 05:20)",
    "timestamp_ms": 违规发生的毫秒时间戳(整数),
    "end_ms": 违规结束的毫秒时间戳(整数),
    "speaker": "说话人标识",
    "original_text": "涉及违规的原始文本内容(原文摘录)",
    "reason": "详细解释为什么违规，必须引用具体规则编号和规则原文",
    "severity": "high 或 medium 或 low",
    "confidence": 0.0到1.0的置信度(浮点数)
}

### 严重程度判定标准
- high: 明确违反禁止性规定（如虚假陈述、承诺收益、同业诋毁、使用禁止字样、不当对比）
- medium: 疑似违规或措辞不当（如夸大但未明确承诺、混淆概念、缺失必要说明）
- low: 轻微不规范（如用词不够严谨、风险提示不充分）"""

_SUMMARY_SYSTEM_PROMPT = """\
你是一个保险行业合规审核专家。请根据给定的违规检查结果，生成一段简明的合规审核总结。

### 要求
1. 概括主要违规类型和数量。
2. 指出最严重的问题。
3. 给出简要的改进建议。
4. 控制在 200 字以内。
5. 不要输出 Markdown 标记，直接输出纯文本。"""


class ComplianceService:
    def __init__(self, client: OllamaClient, settings: Settings) -> None:
        self._client = client
        self._max_text_chars = settings.compliance_max_text_chars
        self._chunk_size = settings.compliance_chunk_size
        self._num_ctx = settings.compliance_num_ctx

    # ------------------------------------------------------------------ #
    #  规则解析
    # ------------------------------------------------------------------ #

    @staticmethod
    def parse_rules(
        file_bytes: bytes, filename: str
    ) -> tuple[list[ComplianceRule], list[str]]:
        """解析规则文件，返回 (规则列表, few-shot 案例列表)。

        支持 CSV 和 XLSX 格式。
        CSV A 列为规则，B-G 列为实际检查结果（作为 few-shot）。
        """
        lower = filename.lower()
        if lower.endswith((".xlsx", ".xls")):
            return _parse_xlsx(file_bytes)
        return _parse_csv(file_bytes)

    # ------------------------------------------------------------------ #
    #  合规审核主入口
    # ------------------------------------------------------------------ #

    async def audit(
        self,
        rules: list[ComplianceRule],
        transcript_entries: list[dict],
        *,
        few_shot_examples: list[str] | None = None,
        on_progress: ProgressCallback | None = None,
    ) -> ComplianceReport:
        """执行合规审核，长文本自动 Map-Reduce。"""
        total_text = sum(len(e.get("text_corrected", "")) for e in transcript_entries)
        if total_text > self._max_text_chars:
            logger.warning(
                "Transcript too long (%d chars), truncating entries", total_text
            )
            truncated: list[dict] = []
            acc = 0
            for e in transcript_entries:
                t = e.get("text_corrected", "")
                if acc + len(t) > self._max_text_chars:
                    break
                truncated.append(e)
                acc += len(t)
            transcript_entries = truncated

        chunks = self._build_entry_chunks(transcript_entries)
        total_steps = len(chunks) + 1  # map chunks + summary
        logger.info(
            "Compliance audit: %d entries -> %d chunks (chunk_size=%d)",
            len(transcript_entries),
            len(chunks),
            self._chunk_size,
        )
        if on_progress:
            on_progress(0, total_steps)

        # Map: 并发审核各 chunk
        completed = 0
        lock = asyncio.Lock()

        async def _audit_with_progress(
            i: int, chunk: list[dict]
        ) -> list[Violation]:
            nonlocal completed
            result = await self._audit_chunk(
                i, len(chunks), rules, chunk, few_shot_examples
            )
            async with lock:
                completed += 1
                if on_progress:
                    on_progress(completed, total_steps)
            return result

        tasks = [
            _audit_with_progress(i, chunk) for i, chunk in enumerate(chunks)
        ]
        chunk_results = await asyncio.gather(*tasks)

        all_violations: list[Violation] = []
        for vs in chunk_results:
            all_violations.extend(vs)

        # 按时间戳排序
        all_violations.sort(key=lambda v: v.timestamp_ms)

        # Reduce: 生成摘要
        summary = await self._generate_summary(rules, all_violations)
        if on_progress:
            on_progress(total_steps, total_steps)

        score = _calculate_score(len(rules), all_violations)

        return ComplianceReport(
            total_rules=len(rules),
            total_segments_checked=len(transcript_entries),
            violations=all_violations,
            summary=summary,
            compliance_score=score,
        )

    # ------------------------------------------------------------------ #
    #  内部方法
    # ------------------------------------------------------------------ #

    def _build_entry_chunks(
        self, entries: list[dict]
    ) -> list[list[dict]]:
        """按字符数将 entries 分组，保证每条 entry 完整保留。"""
        chunks: list[list[dict]] = []
        current: list[dict] = []
        current_len = 0

        for entry in entries:
            text_len = len(entry.get("text_corrected", ""))
            if current and current_len + text_len > self._chunk_size:
                chunks.append(current)
                current = []
                current_len = 0
            current.append(entry)
            current_len += text_len

        if current:
            chunks.append(current)
        return chunks

    async def _audit_chunk(
        self,
        chunk_index: int,
        total_chunks: int,
        rules: list[ComplianceRule],
        entries: list[dict],
        few_shot_examples: list[str] | None = None,
    ) -> list[Violation]:
        """Map 阶段：对单个 chunk 执行 LLM 合规审核。"""
        logger.info(
            "Audit chunk %d/%d (%d entries)...",
            chunk_index + 1,
            total_chunks,
            len(entries),
        )

        # Build timestamp -> precise ms mapping from original entries
        ts_to_ms: dict[str, int] = {}
        ts_to_end_ms: dict[str, int] = {}
        for e in entries:
            ts = e.get("timestamp", "")
            if ts and ts not in ts_to_ms:
                ts_to_ms[ts] = int(e.get("timestamp_ms", 0))
                ts_to_end_ms[ts] = int(e.get("end_ms", 0))

        rules_text = "\n".join(f"{r.id}. {r.content}" for r in rules)
        transcript_lines = [
            f"[{e.get('timestamp', '??:??')}] [{e.get('speaker', '未知')}]: "
            f"{e.get('text_corrected', '')}"
            for e in entries
        ]
        transcript_text = "\n".join(transcript_lines)

        user_parts = [f"【审核标准】\n{rules_text}"]

        if few_shot_examples:
            examples_text = "\n".join(
                f"- {ex}" for ex in few_shot_examples[:5]
            )
            user_parts.append(
                f"【历史违规案例参考】\n{examples_text}\n"
                "（以上为真实违规案例，供你参考判断标准的严格程度。）"
            )

        user_parts.append(
            f"【语音转录文本 - 第 {chunk_index + 1}/{total_chunks} 段】\n"
            f"{transcript_text}"
        )
        user_parts.append(
            "请逐条对照审核标准，仔细检查上述转录文本。\n"
            "注意：\n"
            "1. 对每一条标准都要检查，不要遗漏。\n"
            "2. 包含'不允许出现'或'不得'的规则，只要文本中出现了相应字样（即使有同音字差异），即为违规。\n"
            "3. 将违规原文完整摘录到 original_text 中。\n"
            "4. 有疑似违规的也要报告，severity 标记为 medium。"
        )

        user_prompt = "\n\n".join(user_parts)

        logger.info(
            "Audit chunk %d/%d user prompt (first 1000 chars):\n%s",
            chunk_index + 1,
            total_chunks,
            user_prompt[:1000],
        )

        for attempt in range(1, 3):
            try:
                messages: list[dict[str, str]] = [
                    {"role": "system", "content": _AUDIT_SYSTEM_PROMPT},
                    {"role": "user", "content": user_prompt},
                ]
                if attempt > 1:
                    messages.append(
                        {
                            "role": "user",
                            "content": (
                                "你上次的回答不是合法 JSON 数组。"
                                "请严格只输出 JSON 数组，不要输出任何其他内容。"
                            ),
                        }
                    )

                response = await self._client.chat(
                    messages=messages,
                    json_format=True,
                    num_ctx=self._num_ctx,
                    num_predict=4096,
                    think=False,
                )
                raw = strip_think_tags(response.content)
                logger.info(
                    "Audit chunk %d/%d raw LLM output:\n%s",
                    chunk_index + 1,
                    total_chunks,
                    raw[:2000],
                )
                violations = _parse_violations(raw, rules, ts_to_ms, ts_to_end_ms)
                logger.info(
                    "Audit chunk %d/%d done: %d violations found",
                    chunk_index + 1,
                    total_chunks,
                    len(violations),
                )
                return violations
            except Exception as e:
                logger.warning(
                    "Audit chunk %d/%d attempt %d failed: %s",
                    chunk_index + 1,
                    total_chunks,
                    attempt,
                    e,
                )

        logger.error("Audit chunk %d/%d all attempts failed", chunk_index + 1, total_chunks)
        return []

    async def _generate_summary(
        self,
        rules: list[ComplianceRule],
        violations: list[Violation],
    ) -> str:
        """Reduce 阶段：基于所有违规条目生成整体合规摘要。"""
        if not violations:
            return "审核完成，未发现违规内容。"

        violation_text = "\n".join(
            f"- [{v.timestamp}] [{v.severity}] 违反规则{v.rule_id}: {v.reason}"
            for v in violations
        )
        user_prompt = (
            f"共 {len(rules)} 条审核标准，发现 {len(violations)} 条违规：\n\n"
            f"{violation_text}"
        )

        try:
            response = await self._client.chat(
                messages=[
                    {"role": "system", "content": _SUMMARY_SYSTEM_PROMPT},
                    {"role": "user", "content": user_prompt},
                ],
                num_ctx=self._num_ctx,
                think=False,
                num_predict=1024,
            )
            return strip_think_tags(response.content).strip()
        except Exception as e:
            logger.warning("Summary generation failed: %s", e)
            high = sum(1 for v in violations if v.severity == "high")
            medium = sum(1 for v in violations if v.severity == "medium")
            low = sum(1 for v in violations if v.severity == "low")
            return (
                f"发现 {len(violations)} 条违规"
                f"（高风险 {high} 条，中风险 {medium} 条，低风险 {low} 条）。"
            )


# ------------------------------------------------------------------ #
#  辅助函数
# ------------------------------------------------------------------ #


def _calculate_score(total_rules: int, violations: list[Violation]) -> float:
    """计算合规评分。基础分 100，按 severity 扣分。"""
    deduction = 0.0
    for v in violations:
        if v.severity == "high":
            deduction += 15.0
        elif v.severity == "medium":
            deduction += 8.0
        else:
            deduction += 3.0
    return max(0.0, round(100.0 - deduction, 1))


def _parse_timestamp_to_ms(ts: str) -> int:
    """Parse MM:SS or HH:MM:SS string to milliseconds."""
    parts = ts.strip().split(":")
    try:
        if len(parts) == 2:
            return (int(parts[0]) * 60 + int(parts[1])) * 1000
        if len(parts) == 3:
            return (int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])) * 1000
    except ValueError:
        pass
    return 0


def _safe_int(value: object, default: int = 0) -> int:
    """安全地将 LLM 输出转换为 int，避免非数值字符串崩溃。"""
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_float(value: object, default: float = 0.0) -> float:
    """安全地将 LLM 输出转换为 float。"""
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _parse_violations(
    raw: str,
    rules: list[ComplianceRule],
    ts_to_ms: dict[str, int] | None = None,
    ts_to_end_ms: dict[str, int] | None = None,
) -> list[Violation]:
    """Parse LLM output into Violation list.

    When *ts_to_ms* is provided (timestamp string -> precise ms from original
    transcript entries), LLM-returned ``timestamp_ms`` is replaced with the
    accurate value looked up by the ``timestamp`` string key.
    """
    content = extract_json_array(raw)
    data = json.loads(content)

    # 如果 LLM 输出了对象而非数组
    if isinstance(data, dict):
        # 情况 1: 包装对象 {"violations": [...]}
        for key in ("violations", "results", "items", "data"):
            if key in data and isinstance(data[key], list):
                data = data[key]
                break
        else:
            # 情况 2: 单个 violation 对象 {"rule_id": ..., "reason": ...}
            if "rule_id" in data:
                data = [data]
            else:
                data = []

    if not isinstance(data, list):
        return []

    _VALID_SEVERITY = {"high", "medium", "low"}
    rules_map = {r.id: r.content for r in rules}
    violations: list[Violation] = []

    for item in data:
        if not isinstance(item, dict):
            continue
        rule_id = _safe_int(item.get("rule_id", 0))
        rule_content = item.get("rule_content", "") or rules_map.get(rule_id, "")
        ts_str = str(item.get("timestamp", "00:00"))
        llm_ts_ms = _safe_int(item.get("timestamp_ms", 0))
        llm_end_ms = _safe_int(item.get("end_ms", 0))

        # Validate severity from LLM output
        raw_severity = str(item.get("severity", "low")).lower()
        severity = raw_severity if raw_severity in _VALID_SEVERITY else "low"

        # Resolve precise timestamp_ms from transcript entries mapping.
        # The LLM only sees [MM:SS] strings and cannot reliably infer the
        # original millisecond value, so we look it up from the source data.
        if ts_to_ms and ts_str in ts_to_ms:
            precise_ms = ts_to_ms[ts_str]
            precise_end = ts_to_end_ms.get(ts_str, 0) if ts_to_end_ms else 0
        else:
            # Fallback: parse the timestamp string to approximate ms
            precise_ms = _parse_timestamp_to_ms(ts_str) if not llm_ts_ms else llm_ts_ms
            precise_end = llm_end_ms

        violations.append(
            Violation(
                rule_id=rule_id,
                rule_content=rule_content,
                timestamp=ts_str,
                timestamp_ms=precise_ms,
                end_ms=precise_end if precise_end else precise_ms,
                speaker=str(item.get("speaker", "")),
                original_text=str(item.get("original_text", "")),
                reason=str(item.get("reason", "")),
                severity=severity,
                confidence=_safe_float(item.get("confidence", 0.5), default=0.5),
                source="audio",
            )
        )

    return violations


_HEADER_KEYWORDS = ("必备要素", "检查", "标准", "序号", "注：")
_SKIP_CELLS = {"合格", "不涉及", "None", ""}


def _parse_rule_rows(
    rows: Iterable[list[str]],
) -> tuple[list[ComplianceRule], list[str]]:
    """统一的规则行解析逻辑，CSV 和 XLSX 共用。"""
    rules: list[ComplianceRule] = []
    examples: list[str] = []

    for cells in rows:
        col_a = cells[0].strip() if cells else ""
        if not col_a:
            continue
        if any(kw in col_a for kw in _HEADER_KEYWORDS):
            continue
        if col_a.startswith("存在的问题"):
            break

        rule_id, content = _split_rule_id(col_a, len(rules) + 1)
        if not content:
            continue

        rules.append(ComplianceRule(id=rule_id, content=content))

        for cell in cells[1:]:
            cell = cell.strip()
            if cell and cell not in _SKIP_CELLS:
                examples.append(f"规则{rule_id}({content[:20]}...): {cell}")

    return rules, examples


def _parse_csv(file_bytes: bytes) -> tuple[list[ComplianceRule], list[str]]:
    """解析 CSV 文件，返回 (规则列表, few-shot 案例列表)。"""
    text = _decode_bytes(file_bytes)
    rows = (row for row in csv.reader(io.StringIO(text)) if row)
    return _parse_rule_rows(rows)


def _parse_xlsx(file_bytes: bytes) -> tuple[list[ComplianceRule], list[str]]:
    """解析 XLSX 文件。"""
    try:
        import openpyxl
    except ImportError as e:
        raise ComplianceError("解析 XLSX 需要 openpyxl 库") from e

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        raise ComplianceError("XLSX 文件没有工作表")
    ws = wb[wb.sheetnames[0]]

    rows = (
        [str(c).strip() if c is not None else "" for c in row]
        for row in ws.iter_rows(values_only=True)
    )
    return _parse_rule_rows(rows)


def _decode_bytes(data: bytes) -> str:
    """尝试多种编码解码 CSV。"""
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ComplianceError("无法解码文件，请确保编码为 UTF-8 或 GBK")


def _split_rule_id(text: str, fallback_id: int) -> tuple[int, str]:
    """从 '4全程双录：...' 或 '10资料归档...' 中分离编号和内容。"""
    match = re.match(r"^(\d+)\s*(.+)", text, re.DOTALL)
    if match:
        return int(match.group(1)), match.group(2).strip()
    return fallback_id, text.strip()
